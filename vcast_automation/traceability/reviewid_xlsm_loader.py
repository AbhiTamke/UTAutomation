from __future__ import annotations

from pathlib import Path

from vcast_automation.models.core import TraceabilityRecord


HEADER_ALIASES = {
    "test_case_id": {"testcaseid", "test case id", "test id", "tc id"},
    "unit": {"unit"},
    "function": {"function", "subprogram", "method"},
    "review_id": {"reviewid", "review id", "codebeamer id", "cb id"},
    "requirement_key": {"requirementkey", "requirement key", "requirement"},
    "status": {"status"},
    "asil": {"asil"},
    "notes": {"notes"},
    "description": {"description"},
}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _canonical_header(value: object) -> str | None:
    normalized = _normalize_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


class ReviewIdXlsmLoader:
    def __init__(self, xlsm_path: str | Path | None):
        self.path = Path(xlsm_path) if xlsm_path else None
        self.records: list[TraceabilityRecord] = []
        if self.path and self.path.exists():
            self.records = self._load(self.path)

    def _load(self, path: Path) -> list[TraceabilityRecord]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to read .xlsm ReviewID files.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        output: list[TraceabilityRecord] = []

        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                continue

            header_map: dict[str, int] = {}
            for idx, cell in enumerate(header_row):
                canonical = _canonical_header(cell)
                if canonical:
                    header_map[canonical] = idx

            if "review_id" not in header_map:
                continue

            for row in rows:
                def get(name: str) -> str:
                    idx = header_map.get(name)
                    if idx is None or idx >= len(row):
                        return ""
                    return str(row[idx] or "").strip()

                record = TraceabilityRecord(
                    test_case_id=get("test_case_id"),
                    unit=get("unit"),
                    function=get("function"),
                    subprogram=get("function"),
                    requirement_key=get("requirement_key"),
                    review_id=get("review_id"),
                    status=get("status"),
                    asil=get("asil"),
                    notes=get("notes"),
                    description=get("description"),
                    source=f"{path.name}:{sheet.title}",
                    match_status="loaded",
                    match_confidence="source",
                )

                if record.review_id:
                    output.append(record)

        return output

    def find(self, test_case_id: str, unit: str, function: str, subprogram: str) -> TraceabilityRecord:
        exact_id = [r for r in self.records if r.test_case_id and r.test_case_id == test_case_id]
        if len(exact_id) == 1:
            exact_id[0].match_status = "matched"
            exact_id[0].match_confidence = "high"
            return exact_id[0]
        if len(exact_id) > 1:
            return TraceabilityRecord(test_case_id=test_case_id, unit=unit, function=function, match_status="ambiguous")

        unit_function = [
            r for r in self.records
            if r.unit == unit and (r.function == function or r.subprogram == subprogram)
        ]
        if len(unit_function) == 1:
            unit_function[0].match_status = "matched"
            unit_function[0].match_confidence = "medium"
            return unit_function[0]
        if len(unit_function) > 1:
            return TraceabilityRecord(test_case_id=test_case_id, unit=unit, function=function, match_status="ambiguous")

        return TraceabilityRecord(test_case_id=test_case_id, unit=unit, function=function, match_status="missing")

    def inspect(self) -> dict:
        if not self.path:
            return {"path": "", "exists": False, "records": 0}
        return {
            "path": str(self.path),
            "exists": self.path.exists(),
            "records": len(self.records),
        }